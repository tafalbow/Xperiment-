"""
==============================================================================
INDOEKONOMI data — Indonesia Economic Data Observatory
Export Service: Standardized Multi-Tab Excel (.xlsx) & CSV Generator
==============================================================================
"""

import io
import re
import csv
from datetime import datetime
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExportService:
    """
    Generates production-grade multi-tab Excel files and RFC-4180 CSV exports
    with independent provenance traceability (Sheets: Data, Metadata, Source & Provenance).
    """

    @staticmethod
    def sanitize_filename_part(text: str) -> str:
        """Sanitizes text for file names, eliminating special characters and spaces."""
        text = re.sub(r'[^a-zA-Z0-9_\- ]', '', text)
        text = text.strip().replace(' ', '-')
        return text

    @classmethod
    def generate_filename(cls, dataset_title: str, timeframe: str, extension: str) -> str:
        """
        Creates standardized filename compliant with Section 47:
        INDOEKONOMI_[DATASET TITLE]_[TIMEFRAME].[EXTENSION]
        """
        clean_title = cls.sanitize_filename_part(dataset_title)
        clean_timeframe = cls.sanitize_filename_part(timeframe)
        clean_ext = extension.lstrip('.').lower()
        return f"INDOEKONOMI_{clean_title}_{clean_timeframe}.{clean_ext}"

    @classmethod
    def generate_excel_bytes(
        cls,
        dataset_meta: Dict[str, Any],
        observations: List[Dict[str, Any]],
        provenance_records: List[Dict[str, Any]]
    ) -> bytes:
        """
        Builds a 3-sheet workbook:
        - Sheet 1: Data (with provenance_id foreign key)
        - Sheet 2: Metadata (24-point standard attributes)
        - Sheet 3: Source & Provenance (exact document, page/table, citation)
        """
        wb = openpyxl.Workbook()
        # Default sheet
        ws_data = wb.active
        ws_data.title = "Data Observasi"

        # Styling definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        meta_header_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
        prov_header_fill = PatternFill(start_color="5F6368", end_color="5F6368", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        # ----------------------------------------------------------------------
        # SHEET 1: DATA
        # ----------------------------------------------------------------------
        data_headers = [
            "Observation ID",
            "Provenance ID",
            "Period (Data)",
            "Period Type",
            "Value",
            "Unit",
            "Status",
            "Geography",
            "Source Institution",
            "Publication Title",
            "Publication Date",
            "Page Ref",
            "Table Ref",
            "Extraction Method"
        ]
        ws_data.append(data_headers)

        for col_num in range(1, len(data_headers) + 1):
            cell = ws_data.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for obs in observations:
            prov_id = obs.get("provenance_id") or f"PRV-{obs.get('indicator_id')}-{obs.get('period')}"
            row_data = [
                obs.get("id") or obs.get("observation_id", ""),
                prov_id,
                obs.get("period", ""),
                obs.get("period_type", ""),
                obs.get("value"),
                obs.get("unit", ""),
                obs.get("status", ""),
                obs.get("geography", "Indonesia"),
                obs.get("source_institution") or obs.get("publishing_institution", ""),
                obs.get("publication_title") or obs.get("publication_name", ""),
                obs.get("publication_date", ""),
                obs.get("page_reference") or obs.get("source_page", ""),
                obs.get("table_reference") or obs.get("source_table", ""),
                obs.get("extraction_method") or "Extracted and harmonized by INDOEKONOMI data"
            ]
            ws_data.append(row_data)

        # Apply borders & format numbers
        for row in ws_data.iter_rows(min_row=2, max_row=len(observations) + 1, min_col=1, max_col=len(data_headers)):
            for cell in row:
                cell.border = thin_border
                if cell.column == 5 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

        # Auto column widths
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # ----------------------------------------------------------------------
        # SHEET 2: METADATA
        # ----------------------------------------------------------------------
        ws_meta = wb.create_sheet(title="Metadata")
        meta_headers = ["Atribut Metadata Standar", "Keterangan & Nilai"]
        ws_meta.append(meta_headers)

        for col_num in (1, 2):
            cell = ws_meta.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = meta_header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

        metadata_rows = [
            ("Platform Penerbit", "INDOEKONOMI data — Indonesia Economic Data Observatory"),
            ("Portal Web", "https://indoekonomi.data.go.id"),
            ("Nama Dataset", dataset_meta.get("name") or dataset_meta.get("indicator_name", "")),
            ("Kode Variabel Unik", dataset_meta.get("code") or dataset_meta.get("unique_variable_code", "")),
            ("Sektor", dataset_meta.get("sector", "")),
            ("Kategori", dataset_meta.get("category", "")),
            ("Subkategori", dataset_meta.get("subcategory", "")),
            ("Definisi Operasional", dataset_meta.get("definition", "")),
            ("Satuan Pengukuran", dataset_meta.get("unit", "")),
            ("Frekuensi Pelaporan", dataset_meta.get("frequency", "")),
            ("Cakupan Geografis", dataset_meta.get("geographic_scope", "Indonesia / National")),
            ("Cakupan Periode Data", dataset_meta.get("data_period_coverage", "")),
            ("Metodologi Penghitungan", dataset_meta.get("methodology", "")),
            ("Formula / Rumus Agregasi", dataset_meta.get("calculation_formula", "N/A")),
            ("Lembaga Sumber Resmi", dataset_meta.get("publishing_institution") or dataset_meta.get("source_data", "")),
            ("Dokumen Sumber Kanonikal", dataset_meta.get("publication_document_name", "")),
            ("Kebijakan Status Data", dataset_meta.get("data_status_policy", "")),
            ("URL Sumber Asli", dataset_meta.get("source_url", "")),
            ("Batasan / Catatan Metodologi", dataset_meta.get("data_limitations", "Data resmi hasil harmonisasi statutori")),
            ("Tanggal Ekspor Data", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Status Lisensi & Hak Cipta", "Hak Cipta Pemerintah Republik Indonesia. Lisensi Data Publik Terbuka Terdaftar.")
        ]

        for attr, val in metadata_rows:
            ws_meta.append([attr, val])

        for row in ws_meta.iter_rows(min_row=2, max_row=len(metadata_rows) + 1, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
            row[0].font = Font(name="Calibri", size=11, bold=True)

        ws_meta.column_dimensions['A'].width = 30
        ws_meta.column_dimensions['B'].width = 65

        # ----------------------------------------------------------------------
        # SHEET 3: SOURCE & PROVENANCE
        # ----------------------------------------------------------------------
        ws_prov = wb.create_sheet(title="Source & Provenance")
        prov_headers = [
            "Provenance ID",
            "Institusi Penerbit Resmi",
            "Judul Dokumen Publikasi",
            "Nomor Dokumen Statutori",
            "Tanggal Publikasi",
            "Periode Data",
            "Halaman Sumber",
            "Tabel Sumber",
            "Metode Ekstraksi & Harmonisasi",
            "URL Dokumen Asli"
        ]
        ws_prov.append(prov_headers)

        for col_num in range(1, len(prov_headers) + 1):
            cell = ws_prov.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = prov_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Deduplicate provenance by id
        seen_provs = set()
        for p in provenance_records:
            pid = p.get("provenance_id") or p.get("id")
            if pid in seen_provs:
                continue
            seen_provs.add(pid)
            row_p = [
                pid,
                p.get("source_institution") or p.get("institution_name", ""),
                p.get("publication_title", ""),
                p.get("document_number", "-"),
                p.get("publication_date", ""),
                p.get("edition_period") or p.get("period", ""),
                p.get("page_reference") or "-",
                p.get("table_reference") or "-",
                p.get("extraction_method") or "Extracted and harmonized by INDOEKONOMI data from official source",
                p.get("source_url") or p.get("document_url", "")
            ]
            ws_prov.append(row_p)

        for row in ws_prov.iter_rows(min_row=2, max_row=len(seen_provs) + 1, min_col=1, max_col=len(prov_headers)):
            for cell in row:
                cell.border = thin_border

        for col in ws_prov.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_prov.column_dimensions[col_letter].width = max(max_len + 3, 15)

        # Save to memory stream
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @classmethod
    def generate_csv_bytes(cls, observations: List[Dict[str, Any]]) -> bytes:
        """Builds a standardized RFC-4180 CSV export."""
        output = io.StringIO()
        fieldnames = [
            "observation_id",
            "provenance_id",
            "period",
            "period_type",
            "value",
            "unit",
            "status",
            "geography",
            "source_institution",
            "publication_title",
            "publication_date",
            "page_reference",
            "table_reference",
            "source_document_url"
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()

        for obs in observations:
            prov_id = obs.get("provenance_id") or f"PRV-{obs.get('indicator_id')}-{obs.get('period')}"
            writer.writerow({
                "observation_id": obs.get("id") or obs.get("observation_id", ""),
                "provenance_id": prov_id,
                "period": obs.get("period", ""),
                "period_type": obs.get("period_type", ""),
                "value": obs.get("value"),
                "unit": obs.get("unit", ""),
                "status": obs.get("status", ""),
                "geography": obs.get("geography", "Indonesia"),
                "source_institution": obs.get("source_institution") or obs.get("publishing_institution", ""),
                "publication_title": obs.get("publication_title") or obs.get("publication_name", ""),
                "publication_date": obs.get("publication_date", ""),
                "page_reference": obs.get("page_reference") or obs.get("source_page", ""),
                "table_reference": obs.get("table_reference") or obs.get("source_table", ""),
                "source_document_url": obs.get("document_url") or obs.get("source_url", "")
            })

        return output.getvalue().encode('utf-8')
